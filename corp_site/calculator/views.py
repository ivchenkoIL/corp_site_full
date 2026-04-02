import math

from django.http import HttpResponse
from django.views.generic import FormView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .forms import CableMaterialsForm

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


class CableMaterialsView(FormView):
    template_name = 'calculator/cable_materials.html'
    form_class = CableMaterialsForm

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            form = self.form_class(data=request.GET)
            if form.is_valid():
                results = self.calculate(form.cleaned_data)
                return self.export_excel(form.cleaned_data, results)
        form = self.get_form()
        if form.is_valid():
            results = self.calculate(form.cleaned_data)
            return self.render_to_response(self.get_context_data(form=form, results=results))
        return self.render_to_response(self.get_context_data(form=form))

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        if self.request.method == 'GET' and self.request.GET:
            kwargs['data'] = self.request.GET
        elif self.request.method == 'GET':
            kwargs['data'] = {
                'num_forts': 6,
                'distance': 150,
                'supports_per_span': 1,
                'sip_slack_percent': 5,
                'zoi_per_fort': 2,
                'vvg_length_per_fort': 8,
            }
        return kwargs

    def form_valid(self, form):
        results = self.calculate(form.cleaned_data)
        return self.render_to_response(
            self.get_context_data(form=form, results=results)
        )

    def calculate(self, data):
        num_forts = data['num_forts']
        distance = data['distance']
        supports_per_span = data['supports_per_span']
        sip_slack_percent = data['sip_slack_percent']
        zoi_per_fort = data['zoi_per_fort']
        vvg_length_per_fort = data['vvg_length_per_fort']

        num_spans = num_forts - 1
        sip_length_raw = num_spans * distance
        sip_length = math.ceil(sip_length_raw * (1 + sip_slack_percent / 100))
        total_supports = num_spans * supports_per_span

        if num_forts <= 2:
            anchor_clamps = num_forts
        else:
            anchor_clamps = 2 + (num_forts - 2) * 2

        support_clamps = total_supports
        bracket_l300 = total_supports
        zoi_count = num_forts * zoi_per_fort
        vvg_total = num_forts * vvg_length_per_fort

        # Бандажная лента — фиксированное значение
        band_tape_meters = 50
        # Скрепы для бандажной ленты — фиксированное значение
        band_buckles = 30
        # Изолирующие колпачки — СИП двухжильный: 2 жилы × 2 конца линии
        insulating_caps = 2 * 2

        return {
            'num_spans': num_spans,
            'total_supports': total_supports,
            'sip_length_raw': sip_length_raw,
            'sip_length': sip_length,
            'sip_slack_percent': sip_slack_percent,
            'anchor_clamps': anchor_clamps,
            'support_clamps': support_clamps,
            'bracket_l300': bracket_l300,
            'zoi_count': zoi_count,
            'zoi_per_fort': zoi_per_fort,
            'vvg_total': vvg_total,
            'vvg_length_per_fort': vvg_length_per_fort,
            'band_tape_meters': band_tape_meters,
            'band_buckles': band_buckles,
            'insulating_caps': insulating_caps,
        }

    def export_excel(self, data, results):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Материалы СИП'

        # --- Ширина столбцов ---
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 16

        # --- Стили ---
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        section_font = Font(name='Arial', size=11, bold=True, color='2F5496')
        section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        data_font = Font(name='Arial', size=11)
        bold_font = Font(name='Arial', size=11, bold=True)
        highlight_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center')

        # --- Заголовок ---
        ws.merge_cells('A1:D1')
        cell = ws['A1']
        cell.value = 'Ведомость материалов для прокладки СИП'
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # --- Исходные данные ---
        ws.merge_cells('A3:D3')
        cell = ws['A3']
        cell.value = 'Исходные данные'
        cell.font = section_font
        cell.fill = section_fill
        cell.border = THIN_BORDER
        for c in ['B3', 'C3', 'D3']:
            ws[c].fill = section_fill
            ws[c].border = THIN_BORDER

        params = [
            ('Количество TFortis', f"{data['num_forts']} шт."),
            ('Расстояние между TFortis', f"{data['distance']} м"),
            ('Промежуточных опор в пролёте', f"{data['supports_per_span']} шт."),
            ('Количество пролётов', f"{results['num_spans']} шт."),
            ('Всего промежуточных опор', f"{results['total_supports']} шт."),
            ('Тип СИП', 'двухжильный'),
            ('Запас на провис', f"{results['sip_slack_percent']}%"),
        ]
        for i, (label, value) in enumerate(params, start=4):
            ws.merge_cells(f'A{i}:B{i}')
            c1 = ws[f'A{i}']
            c1.value = label
            c1.font = data_font
            c1.alignment = left
            c1.border = THIN_BORDER
            ws[f'B{i}'].border = THIN_BORDER
            ws.merge_cells(f'C{i}:D{i}')
            c2 = ws[f'C{i}']
            c2.value = value
            c2.font = data_font
            c2.alignment = center
            c2.border = THIN_BORDER
            ws[f'D{i}'].border = THIN_BORDER

        # --- Таблица материалов ---
        table_start = 4 + len(params) + 1
        row = table_start

        # Заголовок таблицы
        headers = ['№', 'Наименование материала', 'Ед. изм.', 'Количество']
        for col_idx, header in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=header)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = THIN_BORDER
        ws.row_dimensions[row].height = 25
        row += 1

        # Данные таблицы
        materials = [
            # (секция, наименование, ед.изм., кол-во, выделить?)
            ('Кабельная продукция', None, None, None, False),
            (None, f'Кабель СИП двухжильный (без запаса: {results["sip_length_raw"]} м)', 'м', results['sip_length'], True),
            (None, 'Кабель ВВГнг 2×2,5', 'м', results['vvg_total'], True),
            ('Арматура для СИП', None, None, None, False),
            (None, 'Анкерные (проходные) зажимы', 'шт.', results['anchor_clamps'], False),
            (None, 'Поддерживающие зажимы', 'шт.', results['support_clamps'], False),
            (None, f'ЗОИ (по {results["zoi_per_fort"]} на TFortis)', 'шт.', results['zoi_count'], False),
            (None, 'Изолирующие колпачки на жилы СИП', 'шт.', results['insulating_caps'], False),
            ('Крепёж и монтаж на опорах', None, None, None, False),
            (None, 'Кронштейн (вылет) L-300', 'шт.', results['bracket_l300'], False),
            (None, 'Бандажная лента', 'м', results['band_tape_meters'], False),
            (None, 'Скрепы для бандажной ленты', 'шт.', results['band_buckles'], False),
        ]

        item_num = 0
        for entry in materials:
            section, name, unit, qty, highlight = entry
            if section:
                # Строка-заголовок секции
                ws.merge_cells(f'A{row}:D{row}')
                c = ws.cell(row=row, column=1, value=section)
                c.font = section_font
                c.fill = section_fill
                c.alignment = left
                c.border = THIN_BORDER
                for col_idx in range(2, 5):
                    ws.cell(row=row, column=col_idx).fill = section_fill
                    ws.cell(row=row, column=col_idx).border = THIN_BORDER
            else:
                item_num += 1
                font = bold_font if highlight else data_font
                fill = highlight_fill if highlight else None

                c1 = ws.cell(row=row, column=1, value=item_num)
                c1.font = font
                c1.alignment = center
                c1.border = THIN_BORDER

                c2 = ws.cell(row=row, column=2, value=name)
                c2.font = font
                c2.alignment = left
                c2.border = THIN_BORDER

                c3 = ws.cell(row=row, column=3, value=unit)
                c3.font = font
                c3.alignment = center
                c3.border = THIN_BORDER

                c4 = ws.cell(row=row, column=4, value=qty)
                c4.font = font
                c4.alignment = center
                c4.border = THIN_BORDER

                if fill:
                    for col_idx in range(1, 5):
                        ws.cell(row=row, column=col_idx).fill = fill

            row += 1

        # --- Примечание ---
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        note = ws.cell(row=row, column=1,
                       value=f'* Запас СИП {results["sip_slack_percent"]}% учитывает провис кабеля между опорами')
        note.font = Font(name='Arial', size=10, italic=True, color='666666')
        note.alignment = left

        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        note2 = ws.cell(row=row, column=1,
                        value=f'* ВВГнг 2×2,5 — по {results["vvg_length_per_fort"]} м на каждый TFortis')
        note2.font = Font(name='Arial', size=10, italic=True, color='666666')
        note2.alignment = left

        # --- Печать ---
        ws.print_area = f'A1:D{row}'
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="materials_sip.xlsx"'
        wb.save(response)
        return response
